"""
PA-Shift Generator Backend Server
proto-type.py をベースに Flask で API サーバーを構築
複数日対応版
"""

from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from datetime import datetime, timedelta
import copy
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO

app = Flask(__name__)
CORS(app)  # CORS対応

DESK_SCORE_THRESHOLD = 5
STAGE_SCORE_THRESHOLD = 3
CONTINUITY_BONUS = 20


def parse_time_slot(time_slot):
    """
    "HH:MM-HH:MM" 形式を datetime の開始/終了に変換
    """
    start_str, end_str = time_slot.split("-")
    start_dt = datetime.strptime(start_str, "%H:%M")
    end_dt = datetime.strptime(end_str, "%H:%M")
    return start_dt, end_dt


def has_time_overlap(slot_a, slot_b):
    """
    2つの時間帯が1分でも重なれば True
    """
    start_a, end_a = parse_time_slot(slot_a)
    start_b, end_b = parse_time_slot(slot_b)
    return start_a < end_b and start_b < end_a


def member_ng_times_for_day(member, day_num=None):
    """
    メンバーのNG時間を日別に取り出す（非破壊）
    """
    ng_times = member.get("ng_times", [])
    if isinstance(ng_times, dict):
        if day_num is None:
            return []
        return ng_times.get(f"day_{day_num}", [])
    if isinstance(ng_times, list):
        return ng_times
    return []


def has_time_conflict(band_slots, member, day_num=None):
    """
    バンド時間帯とメンバーNG時間帯が重複するか
    """
    member_ng_slots = member_ng_times_for_day(member, day_num)
    for band_slot in band_slots:
        for ng_slot in member_ng_slots:
            if has_time_overlap(band_slot, ng_slot):
                return True
    return False


def validate_member_structure(member, index):
    """
    メンバー1件の必須項目と型・値を検証する
    """
    required_keys = ["name", "skill_desk", "skill_stage", "count", "ng_bands"]
    for key in required_keys:
        if key not in member:
            return f"members[{index}] に必須キー '{key}' がありません"

    if not isinstance(member["name"], str) or not member["name"].strip():
        return f"members[{index}].name は空でない文字列である必要があります"

    numeric_fields = ["skill_desk", "skill_stage", "count"]
    for field in numeric_fields:
        if not isinstance(member[field], (int, float)):
            return f"members[{index}].{field} は数値である必要があります"
        if member[field] < 0:
            return f"members[{index}].{field} は0以上である必要があります"

    if not isinstance(member["ng_bands"], list):
        return f"members[{index}].ng_bands は配列である必要があります"

    req_bands = member.get("req_bands", [])
    if req_bands is not None and not isinstance(req_bands, list):
        return f"members[{index}].req_bands は配列である必要があります"

    ng_times = member.get("ng_times", [])
    if isinstance(ng_times, list):
        for i, slot in enumerate(ng_times):
            if not isinstance(slot, str):
                return f"members[{index}].ng_times[{i}] は文字列である必要があります"
            try:
                parse_time_slot(slot)
            except (ValueError, TypeError):
                return f"members[{index}].ng_times[{i}] は 'HH:MM-HH:MM' 形式である必要があります"
    elif isinstance(ng_times, dict):
        for day_key, slots in ng_times.items():
            if not isinstance(day_key, str) or not day_key.startswith("day_"):
                return f"members[{index}].ng_times のキーは 'day_n' 形式である必要があります"
            if not isinstance(slots, list):
                return f"members[{index}].ng_times['{day_key}'] は配列である必要があります"
            for i, slot in enumerate(slots):
                if not isinstance(slot, str):
                    return f"members[{index}].ng_times['{day_key}'][{i}] は文字列である必要があります"
                try:
                    parse_time_slot(slot)
                except (ValueError, TypeError):
                    return f"members[{index}].ng_times['{day_key}'][{i}] は 'HH:MM-HH:MM' 形式である必要があります"
    else:
        return f"members[{index}].ng_times は配列または日別オブジェクトである必要があります"

    return None


def validate_members(members):
    """
    メンバー配列全体の検証
    """
    if not isinstance(members, list) or len(members) == 0:
        return "メンバーが1人以上必要です"

    for index, member in enumerate(members):
        if not isinstance(member, dict):
            return f"members[{index}] はオブジェクトである必要があります"
        error = validate_member_structure(member, index)
        if error:
            return error
    return None


def day_sort_key(day_key):
    """
    day_1, day_2, ... を数値としてソートする
    """
    try:
        return int(str(day_key).split("_")[1])
    except (IndexError, ValueError, TypeError):
        return float("inf")


def generate_timetable(start_time_str, band_list, rh_mins, act_mins, break_info=None):
    """
    開始時間とバンドリストから、タイムテーブルを自動生成する関数
    """
    timetable = []
    # 文字列の時間を、計算できる「時計データ」に変換
    current_time = datetime.strptime(start_time_str, "%H:%M")

    for band in band_list:
        # 1. リハの時間を計算して追加
        rh_start = current_time.strftime("%H:%M")
        current_time += timedelta(minutes=rh_mins)  # リハの分数だけ時間を進める
        rh_end = current_time.strftime("%H:%M")
        timetable.append({"time": f"{rh_start}-{rh_end}", "type": "rh", "name": band})

        # 2. 本番の時間を計算して追加
        act_start = current_time.strftime("%H:%M")
        current_time += timedelta(minutes=act_mins)  # 本番の分数だけ時間を進める
        act_end = current_time.strftime("%H:%M")
        timetable.append({"time": f"{act_start}-{act_end}", "type": "act", "name": band})

        # 3. お昼休みの判定（もし指定されていて、今のバンドが終わった直後なら）
        if break_info and break_info.get("after_band") == band:
            break_start = current_time.strftime("%H:%M")
            current_time += timedelta(minutes=break_info["duration"])  # 休憩の分数だけ進める
            break_end = current_time.strftime("%H:%M")
            timetable.append({"time": f"{break_start}-{break_end}", "type": "break", "name": "昼休憩"})

    return timetable


def generate_pa_shift(timetable, members_data, day_num=None):
    """
    PAシフトを作成する関数（v2：リハ・本番セット化＆インターバル制約対応）
    """
    members = copy.deepcopy(members_data)
    shift_result = {}
    infeasible_bands = []

    # 1. タイムテーブルから「シフト対象のバンド」と「前後の順番」を整理する
    band_times = {}  # {"バンド名": ["リハ時間", "本番時間"]}
    band_order = []  # 前後関係を把握するためのリスト

    for entry in timetable:
        if entry["type"] == "break":
            continue  # 休憩はシフト計算から除外

        b_name = entry["name"]
        if b_name not in band_times:
            band_times[b_name] = []
            band_order.append(b_name)
        band_times[b_name].append(entry["time"])

    # 各メンバーの仕事回数の上限を計算（バンド数*4/登録メンバー数）
    num_bands = len(band_order)
    num_members = len(members)
    shift_limit = (num_bands * 4) / num_members if num_members > 0 else float('inf')

    # 2. バンドごとにシフトを計算（ここでリハと本番が1セットとして扱われます）
    for i, band in enumerate(band_order):
        desk_team = []
        stage_team = []
        desk_score = 0
        stage_score = 0

        # --- 新仕様：前後のバンドを取得 ---
        prev_band = band_order[i - 1] if i > 0 else None
        next_band = band_order[i + 1] if i < len(band_order) - 1 else None

        # 前のバンドで割り当てられた人を取得（連続性ボーナス用）
        prev_assigned = set()
        if prev_band and prev_band in shift_result:
            prev_assigned = set(shift_result[prev_band]["卓"] + shift_result[prev_band]["ステージ"])

        # 卓チーム編成
        available_desk = []
        for m in members:
            # 上限チェック：仕事回数が上限に達していないか
            if m["count"] >= shift_limit:
                continue

            # NG判定①：自分が出演するバンド、またはその「前後」ならシフト不可
            if (
                band in m["ng_bands"]
                or prev_band in m["ng_bands"]
                or next_band in m["ng_bands"]
            ):
                continue

            # NG判定②：LINEで指定されたNG時間に被っているか
            if has_time_conflict(band_times[band], m, day_num):
                continue

            # 卓優先度計算
            priority_desk = -m["count"] * 10
            if band in m.get("req_bands", []):
                priority_desk += 100
            priority_desk += m["skill_desk"]
            
            # 連続性ボーナス：前のバンドで割り当てられた人を優先
            if m["name"] in prev_assigned:
                priority_desk += CONTINUITY_BONUS

            candidate = m.copy()
            candidate["priority_desk"] = priority_desk
            available_desk.append(candidate)

        available_desk.sort(key=lambda x: x["priority_desk"], reverse=True)

        for m in available_desk:
            if desk_score < DESK_SCORE_THRESHOLD:
                desk_team.append(m)
                desk_score += m["skill_desk"]

        # ステージチーム編成
        available_stage = []
        desk_team_names = [m["name"] for m in desk_team]
        for m in members:
            if m["name"] in desk_team_names:  # すでに卓に割り当てられたメンバーは除く
                continue
            
            # 上限チェック：仕事回数が上限に達していないか
            if m["count"] >= shift_limit:
                continue

            # NG判定①：自分が出演するバンド、またはその「前後」ならシフト不可
            if (
                band in m["ng_bands"]
                or prev_band in m["ng_bands"]
                or next_band in m["ng_bands"]
            ):
                continue

            # NG判定②：LINEで指定されたNG時間に被っているか
            if has_time_conflict(band_times[band], m, day_num):
                continue

            # ステージ優先度計算
            priority_stage = -m["count"] * 10
            if band in m.get("req_bands", []):
                priority_stage += 100
            priority_stage += m["skill_stage"]
            
            # 連続性ボーナス：前のバンドで割り当てられた人を優先
            if m["name"] in prev_assigned:
                priority_stage += CONTINUITY_BONUS

            candidate = m.copy()
            candidate["priority_stage"] = priority_stage
            available_stage.append(candidate)

        available_stage.sort(key=lambda x: x["priority_stage"], reverse=True)

        for m in available_stage:
            if stage_score < STAGE_SCORE_THRESHOLD:
                stage_team.append(m)
                stage_score += m["skill_stage"]

        # 3. 結果の保存とカウント更新（2枠セットで1カウント）
        shift_result[band] = {
            "卓": [m["name"] for m in desk_team],
            "ステージ": [m["name"] for m in stage_team],
        }

        assigned_names = [m["name"] for m in desk_team + stage_team]
        for m in members:
            if m["name"] in assigned_names:
                m["count"] += 1

        if desk_score < DESK_SCORE_THRESHOLD or stage_score < STAGE_SCORE_THRESHOLD:
            infeasible_bands.append(
                {
                    "band": band,
                    "desk_score": desk_score,
                    "stage_score": stage_score,
                    "desk_required": DESK_SCORE_THRESHOLD,
                    "stage_required": STAGE_SCORE_THRESHOLD,
                }
            )

    return shift_result, members, infeasible_bands


def generate_timetable_multi_day(timetable_config):
    """
    複数日のタイムテーブルを生成する関数
    timetable_config: {
        "num_days": 2,
        "days": [
            {
                "day_number": 1,
                "start_time": "11:30",
                "bands": ["band1", "band2"],
                "rh_mins": 15,
                "act_mins": 10,
                "break_duration": 60,
                "break_after_band": "band1"
            },
            {...}
        ]
    }
    """
    result = {}
    for day_config in timetable_config.get("days", []):
        day_num = day_config.get("day_number", 1)
        timetable = generate_timetable(
            day_config.get("start_time", "11:30"),
            day_config.get("bands", []),
            day_config.get("rh_mins", 15),
            day_config.get("act_mins", 10),
            {
                "after_band": day_config.get("break_after_band", ""),
                "duration": day_config.get("break_duration", 60)
            } if day_config.get("break_after_band") else None
        )
        result[f"day_{day_num}"] = timetable
    return result


def generate_pa_shift_multi_day(timetable_multi, members_data):
    """
    複数日のシフトを生成する関数
    """
    members = copy.deepcopy(members_data)
    shift_result = {}
    infeasible_days = {}

    for day_key, timetable in sorted(timetable_multi.items(), key=lambda item: day_sort_key(item[0])):
        # 日番号を抽出
        day_num = day_sort_key(day_key)

        # その日のシフトを生成（ng_timesは破壊的に上書きしない）
        day_shift, members, infeasible_bands = generate_pa_shift(timetable, members, day_num=day_num)
        shift_result[day_key] = day_shift
        if infeasible_bands:
            infeasible_days[day_key] = infeasible_bands

    return shift_result, members, infeasible_days


def create_excel_workbook(timetable_multi, shift_result, members):
    """
    複数日のシフトをエクセルワークブックとして作成
    各日ごとに1シート、最後に集計シートを追加
    """
    wb = Workbook()
    wb.remove(wb.active)  # デフォルトシートを削除

    # スタイル定義
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 各日のシートを作成
    day_sheets = {}
    for day_key, timetable in sorted(timetable_multi.items(), key=lambda item: day_sort_key(item[0])):
        day_num = int(day_key.split('_')[1])
        sheet_name = f"{day_num}日目"
        ws = wb.create_sheet(sheet_name)
        day_sheets[day_key] = ws

        # ヘッダー行
        headers = ["時間帯", "種別", "バンド名", "担当（卓）", "担当（ステージ）"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # データ行
        for row, entry in enumerate(timetable, 2):
            assigned = shift_result[day_key].get(entry["name"], {"卓": [], "ステージ": []})
            desk = "、".join(assigned["卓"]) if assigned["卓"] else "-"
            stage = "、".join(assigned["ステージ"]) if assigned["ステージ"] else "-"
            entry_type = "リハ" if entry["type"] == "rh" else "本番" if entry["type"] == "act" else "休憩"

            row_data = [entry["time"], entry_type, entry["name"], desk, stage]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center")

        # 列幅を調整
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20

    # 全体集計シートを作成
    ws_summary = wb.create_sheet("全体集計", 0)

    # 集計ヘッダー
    summary_headers = ["メンバー名", "シフト回数"]
    for col, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # メンバーをカウント値で降順ソート
    sorted_members = sorted(members, key=lambda x: x["count"], reverse=True)

    # データ行
    for row, member in enumerate(sorted_members, 2):
        member_data = [member["name"], member["count"]]
        for col, value in enumerate(member_data, 1):
            cell = ws_summary.cell(row=row, column=col)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")

    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 15

    return wb


def check_ng_time_for_day(ng_times_per_day, day_num, time_slot):
    """
    指定された日のNG時間内にタイムスロットが含まれているかチェック
    ng_times_per_day: {"day_1": ["11:30-12:00", "13:00-14:00"], ...}
    day_num: 日番号（1, 2, ...）
    time_slot: 時間スロット（"11:30-12:00"）
    """
    day_key = f"day_{day_num}"
    if day_key not in ng_times_per_day:
        return False
    return any(has_time_overlap(time_slot, ng_slot) for ng_slot in ng_times_per_day[day_key])




@app.route("/api/generate-shift-multi-day", methods=["POST"])
def api_generate_shift_multi_day():
    """
    複数日のタイムテーブルとシフトを生成する統合エンドポイント
    """
    try:
        data = request.json

        # リクエストデータの取得
        num_days = data.get("num_days", 1)
        days = data.get("days", [])
        members = data.get("members", [])

        # バリデーション
        if num_days < 1:
            return jsonify({"error": "イベント日数は1日以上である必要があります"}), 400
        if not days or len(days) == 0:
            return jsonify({"error": "各日のバンド設定が必要です"}), 400
        members_error = validate_members(members)
        if members_error:
            return jsonify({"error": members_error}), 400

        # 複数日タイムテーブル生成
        timetable_multi = generate_timetable_multi_day({"num_days": num_days, "days": days})

        # 複数日シフト生成
        shift_result, updated_members, infeasible_days = generate_pa_shift_multi_day(timetable_multi, members)
        if infeasible_days:
            return (
                jsonify(
                    {
                        "error": "条件を満たすシフトを生成できませんでした",
                        "infeasible_days": infeasible_days,
                    }
                ),
                400,
            )

        # レスポンス作成
        return jsonify(
            {
                "status": "success",
                "timetable_multi": timetable_multi,
                "shift": shift_result,
                "members": updated_members,
            }
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/download-excel", methods=["POST"])
def api_download_excel():
    """
    複数日のシフトをエクセルファイルとしてダウンロード
    """
    try:
        data = request.json

        # リクエストデータの取得
        timetable_multi = data.get("timetable_multi", {})
        shift = data.get("shift", {})
        members = data.get("members", [])

        # バリデーション
        if not timetable_multi or not shift:
            return jsonify({"error": "シフトデータが必要です"}), 400

        # エクセルワークブック作成
        wb = create_excel_workbook(timetable_multi, shift, members)

        # BytesIOに出力
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # ファイルとして返す
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"pa_shift_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/generate-shift", methods=["POST"])
def api_generate_shift():
    """
    タイムテーブル生成とシフト生成を行う統合エンドポイント
    """
    try:
        data = request.json

        # リクエストデータの取得
        start_time = data.get("start_time", "11:30")
        bands = data.get("bands", [])
        rh_mins = data.get("rh_mins", 15)
        act_mins = data.get("act_mins", 10)
        break_duration = data.get("break_duration", 60)
        break_after_band = data.get("break_after_band", "")
        members = data.get("members", [])

        # バリデーション
        if not bands:
            return jsonify({"error": "バンドが1つ以上必要です"}), 400
        members_error = validate_members(members)
        if members_error:
            return jsonify({"error": members_error}), 400

        # breakInfoの組み立て
        break_info = None
        if break_after_band:
            break_info = {"after_band": break_after_band, "duration": break_duration}

        # タイムテーブル生成
        timetable = generate_timetable(start_time, bands, rh_mins, act_mins, break_info)

        # シフト生成
        shift_result, updated_members, infeasible_bands = generate_pa_shift(timetable, members)
        if infeasible_bands:
            return (
                jsonify(
                    {
                        "error": "条件を満たすシフトを生成できませんでした",
                        "infeasible_bands": infeasible_bands,
                    }
                ),
                400,
            )

        # レスポンス作成
        return jsonify(
            {
                "status": "success",
                "timetable": timetable,
                "shift": shift_result,
                "members": updated_members,
            }
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/health", methods=["GET"])
def health_check():
    """ヘルスチェック"""
    return jsonify({"status": "ok", "message": "PA-Shift Backend is running"})


if __name__ == "__main__":
    print("🚀 PA-Shift Backend Server starting...")
    print("📌 http://localhost:5000 でサーバーが起動しました")
    print("🔗 http://localhost:5000/api/health でヘルスチェック")
    app.run(debug=True, port=5000)
